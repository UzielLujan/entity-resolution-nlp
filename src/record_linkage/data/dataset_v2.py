"""Dataset v2: two-step labeling pipeline that produces entity_id with classified pairs."""

from pathlib import Path
from typing import List, Optional, Union

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from record_linkage.data.serialization import serialize_record, serialize_record_zeroshot
from record_linkage.utils.normalization import normalizar_nombre_v2
from record_linkage.utils.pairs import _COL_MAP, build_pairs_df, classify_pairs

POSITIVE_CRITERIA = {"llave_exacta", "metrica_clasica"}

# Orden canónico: econo → comor → ts. Debe coincidir con el orden en que se pasan csv_paths.
_DEFAULT_SOURCE_NAMES = ["Económico", "Comorbilidad", "Trabajo Social"]
_CSV_TYPES = ["econo", "comorbilidad", "trabajo_social"]

# Paleta de estilos del xlsx de revisión (tonos pastel suaves, estilo "bootstrap alert").
_FILL_HEADER         = PatternFill("solid", fgColor="9DC7ED")  # gris azul cabecera
_FILL_MATCH          = PatternFill("solid", fgColor="D4EDDA")  # verde:    par positivo (decision=match o auto-confirmado)
_FILL_NO_MATCH       = PatternFill("solid", fgColor="F8D7DA")  # rojo:     par negativo (decision=no_match)
_FILL_NO_CONFIRMADO  = PatternFill("solid", fgColor="FFF3CD")  # amarillo: pendiente de revisión manual
_ALIGN_CENTER        = Alignment(horizontal="center", vertical="center")
_FONT_HEADER         = Font(bold=True)
_MAX_COL_WIDTH       = 50  # tope para columnas de texto largo


def _union_find_entity_ids(n_records: int, positive_pairs: list) -> list:
    """Asigna entity_ids via union-find sobre pares positivos.

    Registros conectados por pares positivos reciben el mismo entity_id.
    Singletons (sin par positivo) reciben un entity_id propio.
    """
    parent = list(range(n_records))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for a, b in positive_pairs:
        union(int(a), int(b))

    root_to_id: dict = {}
    entity_ids = []
    for i in range(n_records):
        root = find(i)
        if root not in root_to_id:
            root_to_id[root] = len(root_to_id)
        entity_ids.append(root_to_id[root])

    return entity_ids


def _pairs_to_review_df(pairs_classified: pd.DataFrame) -> pd.DataFrame:
    """Proyecta pairs_classified a las columnas visibles del xlsx editable.

    Columnas resultantes (en orden):
        record_id_a, record_id_b         → llaves técnicas (se ocultan visualmente)
        source_a, source_b               → identificación del cruce
        exp                              → expediente compartido (vacío para pares NaN-TS)
        nombre_norm_a, nombre_norm_b     → nombres normalizados (lo que jw/lev comparan)
        jw, lev                          → métricas redondeadas a 3 decimales
        criterio                         → clasificación del pipeline (auditoría)
        decision                         → vacío inicialmente; el revisor llena con match/no_match
    """
    return pd.DataFrame({
        "record_id_a":   pairs_classified["record_id_a"],
        "record_id_b":   pairs_classified["record_id_b"],
        "source_a":      pairs_classified["source_a"],
        "source_b":      pairs_classified["source_b"],
        "exp":           pairs_classified["exp_a"],  # exp_a == exp_b en cross-CSV; NA para NaN-TS
        "nombre_norm_a": pairs_classified["nombre_norm_a"],
        "nombre_norm_b": pairs_classified["nombre_norm_b"],
        "jw":            pairs_classified["jw_score"].round(3),
        "lev":           pairs_classified["lev_score"].round(3),
        "criterio":      pairs_classified["criterio"],
        "decision":      pd.Series([None] * len(pairs_classified), dtype="object"),
    })


def _write_review_xlsx(review_df: pd.DataFrame, xlsx_path: Path) -> None:
    """Escribe xlsx con styling, dropdown en 'decision' y record_id_a/b ocultos.

    Aplica en cada escritura:
      - cabecera gris azulada + negrita + congelada (freeze_panes A2)
      - centrado de todos los valores
      - anchos de columna auto-ajustados (con tope _MAX_COL_WIDTH)
      - relleno por fila según 'criterio': verde para auto-confirmados
        (llave_exacta, metrica_clasica), amarillo para no_confirmado.
        La columna 'decision' queda blanca en todas las filas para destacar
        visualmente que es la única celda editable.
      - dropdown match/no_match en la columna 'decision'

    Todo se reaplica en cada escritura porque pandas+openpyxl no preserva
    validaciones, fills ni alignment custom en un round-trip de lectura/escritura.
    """
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="pairs", index=False)
        ws = writer.sheets["pairs"]

        cols = list(review_df.columns)
        n_rows = len(review_df)
        n_cols = len(cols)
        decision_idx = cols.index("decision") + 1
        criterio_idx = cols.index("criterio") + 1
        decision_letter = get_column_letter(decision_idx)

        # record_id_a / record_id_b se dejan visibles por defecto — útiles para
        # extraer pares con scripts/show_pair.py durante revisión manual.
        # Ocultarlas manualmente desde LibreOffice/Excel si estorban visualmente.

        # 1) Cabecera: fondo gris azulado, negrita, centrada
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER
            cell.alignment = _ALIGN_CENTER

        # 2) Centrado de valores + fill por fila según 'decision' (prioridad) o 'criterio'.
        #    Reglas:
        #      decision=match                                              → verde
        #      decision=no_match                                           → rojo
        #      decision vacío + criterio en {llave_exacta, metrica_clasica} → verde (post-classify)
        #      decision vacío + criterio=no_confirmado                      → amarillo (pendiente)
        #    La columna 'decision' nunca recibe fill — queda blanca para destacar como editable.
        for row_idx in range(2, n_rows + 2):
            decision_val = ws.cell(row=row_idx, column=decision_idx).value
            criterio_val = ws.cell(row=row_idx, column=criterio_idx).value
            decision_empty = decision_val is None or decision_val == ""

            if decision_val == "match":
                row_fill = _FILL_MATCH
            elif decision_val == "no_match":
                row_fill = _FILL_NO_MATCH
            elif decision_empty and criterio_val in ("llave_exacta", "metrica_clasica"):
                row_fill = _FILL_MATCH
            elif decision_empty and criterio_val == "no_confirmado":
                row_fill = _FILL_NO_CONFIRMADO
            else:
                row_fill = None

            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = _ALIGN_CENTER
                if row_fill is not None and col_idx != decision_idx:
                    cell.fill = row_fill

        # 3) Anchos de columna auto-ajustados (header vs contenido, con tope)
        for col_idx in range(1, n_cols + 1):
            col_letter = get_column_letter(col_idx)
            header_len = len(str(cols[col_idx - 1]))
            content_lengths = (
                len(str(v)) if not pd.isna(v) else 0
                for v in review_df.iloc[:, col_idx - 1]
            )
            max_len = max(header_len, max(content_lengths, default=0))
            ws.column_dimensions[col_letter].width = min(max_len + 2, _MAX_COL_WIDTH)

        # 4) Congelar la cabecera (siempre visible al hacer scroll vertical)
        ws.freeze_panes = "A2"

        # 5) Dropdown match/no_match en la columna 'decision'
        dv = DataValidation(type="list", formula1='"match,no_match"', allow_blank=True)
        dv.error = "Solo se permite 'match' o 'no_match'"
        dv.errorTitle = "Valor inválido"
        dv.prompt = "Selecciona match o no_match (o deja vacío para usar el criterio del pipeline)"
        dv.promptTitle = "Decisión manual"
        ws.add_data_validation(dv)
        dv.add(f"{decision_letter}2:{decision_letter}{n_rows + 1}")


def build_dataset_v2(
    csv_paths: List,
    output_dir: Union[str, Path],
    source_db_names: Optional[List[str]] = None,
    use_block_tokens: bool = True,
    step: str = "classify",
    umbral_jw: float = 0.88,
    umbral_lev: float = 0.85,
) -> pd.DataFrame:
    """Pipeline de etiquetado v2 — orquestador de dos pasos.

    Args:
        csv_paths:        3 rutas a CSVs limpios en orden fijo [econo, comor, ts]
        output_dir:       Directorio de artefactos intermedios y finales
        source_db_names:  Nombres de source_db (default: ['Económico', 'Comorbilidad', 'Trabajo Social'])
        use_block_tokens: True → tokens [BLK_*] para entrenamiento; False → texto limpio (zero-shot)
        step:             'classify' → produce records_interim.parquet + pairs_classified.parquet
                                       + pairs_for_review.xlsx (editable, con dropdown en 'decision')
                          'finalize' → lee pairs_for_review.xlsx editado, aplica decisiones,
                                       produce dataset_v2.parquet y regenera el xlsx con entity_id
        umbral_jw:        Umbral Jaro-Winkler para metrica_clasica (calibrado: 0.88)
        umbral_lev:       Umbral Levenshtein para metrica_clasica (calibrado: 0.85)

    Returns:
        step='classify': pairs_classified DataFrame
        step='finalize': dataset_v2 DataFrame (record_id, source_db, text, entity_id)
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "interim").mkdir(exist_ok=True)

    if source_db_names is None:
        source_db_names = _DEFAULT_SOURCE_NAMES

    if step == "classify":
        return _step_classify(
            csv_paths, output_dir, source_db_names, use_block_tokens,
            umbral_jw, umbral_lev,
        )
    elif step == "finalize":
        # Legacy: paths derivados de output_dir; text pre-existente en records_interim (no re-serializa).
        return _step_finalize(
            records_path=output_dir / "interim" / "records_interim.parquet",
            xlsx_path=output_dir / "pairs_for_review.xlsx",
            output_path=output_dir / "dataset_v2.parquet",
            csv_paths=None,
            use_block_tokens=use_block_tokens,
            skip_null=False,
        )
    else:
        raise ValueError(f"step '{step}' no reconocido. Usar 'classify' o 'finalize'.")


def _step_classify(
    csv_paths, output_dir, source_db_names, use_block_tokens,
    umbral_jw, umbral_lev,
) -> pd.DataFrame:
    """Lee CSVs → serializa → construye pares → clasifica → escribe intermedios + xlsx editable."""
    csv_paths = [Path(p) for p in csv_paths]
    if len(csv_paths) != 3:
        raise ValueError("csv_paths debe contener exactamente 3 rutas en orden [econo, comor, ts]")

    dfs = []
    for path in csv_paths:
        if not path.exists():
            raise FileNotFoundError(f"CSV no encontrado: {path}")
        dfs.append(pd.read_csv(path))

    df_econo, df_comor, df_ts = dfs

    # Serializar y asignar record_ids en orden canónico: econo → comor → ts
    # INVARIANTE: mismo orden que build_pairs_df usa internamente (Económico → Comorbilidad → Trabajo Social)
    records_rows = []
    record_id = 0
    for df, source, csv_type in zip(dfs, source_db_names, _CSV_TYPES):
        exp_col, nombre_col = _COL_MAP[csv_type]
        for _, row in df.iterrows():
            if use_block_tokens:
                text = serialize_record(row, csv_name=csv_type, use_block_tokens=True)
            else:
                text = serialize_record_zeroshot(row)
            exp_raw = row.get(exp_col)
            records_rows.append({
                "record_id":   record_id,
                "source_db":   source,
                "text":        text,
                "exp_int":     pd.to_numeric(exp_raw, errors="coerce"),
                "nombre_norm": normalizar_nombre_v2(row.get(nombre_col, "")),
            })
            record_id += 1

    interim_dir = output_dir / "interim"

    records_df = pd.DataFrame(records_rows)
    records_path = interim_dir / "records_interim.parquet"
    records_df.to_parquet(records_path, engine="pyarrow", index=False, compression="snappy")
    print(f"✓ interim/records_interim.parquet: {len(records_df):,} registros")

    # Construir y clasificar pares
    pairs_df = build_pairs_df(df_econo, df_comor, df_ts)
    pairs_classified = classify_pairs(
        pairs_df,
        umbral_jw=umbral_jw,
        umbral_lev=umbral_lev,
    )

    pairs_path = interim_dir / "pairs_classified.parquet"
    pairs_classified.to_parquet(pairs_path, engine="pyarrow", index=False, compression="snappy")

    # xlsx editable para revisión manual — superficie de edición de 'decision'
    review_df = _pairs_to_review_df(pairs_classified)
    xlsx_path = output_dir / "pairs_for_review.xlsx"
    _write_review_xlsx(review_df, xlsx_path)

    _print_classify_summary(pairs_classified)
    print(f"✓ interim/pairs_classified.parquet: {len(pairs_classified):,} pares")
    print(f"✓ pairs_for_review.xlsx (editable): {xlsx_path}")
    print(f"\n→ Revisar 'no_confirmado' en pairs_for_review.xlsx (columna 'decision': match/no_match)")
    print(f"  Las celdas vacías de 'decision' usan el criterio del pipeline.")
    print(f"  Luego corre --step finalize.")

    return pairs_classified


def _step_finalize(
    records_path: Path,
    xlsx_path: Path,
    output_path: Path,
    csv_paths: Optional[List[Path]] = None,
    use_block_tokens: bool = True,
    skip_null: bool = False,
) -> pd.DataFrame:
    """Lee xlsx editado → aplica decisiones → union-find entity_ids → escribe parquet + regenera xlsx.

    Args:
        records_path:     interim/records_interim.parquet (record_id, source_db, exp_int, nombre_norm)
        xlsx_path:        pairs_for_review.xlsx con decisiones manuales
        output_path:      ruta completa al parquet final (incluye nombre, no es directorio)
        csv_paths:        si se provee, re-serializa text desde los CSVs limpios con la config indicada.
                          Si None, usa la columna text pre-existente en records_path (comportamiento legacy).
        use_block_tokens: aplica solo si csv_paths se provee
        skip_null:        aplica solo si csv_paths se provee
    """
    for path in [records_path, xlsx_path]:
        if not path.exists():
            raise FileNotFoundError(
                f"Archivo no encontrado: {path}\n"
                "Ejecutar primero con --step classify"
            )

    records_df = pd.read_parquet(records_path)
    review_df  = pd.read_excel(xlsx_path, sheet_name="pairs", engine="openpyxl")

    # Re-serialización opcional: cuando csv_paths está dado, sobrescribimos la columna `text`
    # iterando los CSVs en el mismo orden que _step_classify (econo → comor → ts).
    # INVARIANTE: el record_id es asignado linealmente en ese orden — re-iterar produce el mismo mapping.
    if csv_paths is not None:
        if len(csv_paths) != 3:
            raise ValueError("csv_paths debe contener exactamente 3 rutas en orden [econo, comor, ts]")
        new_texts = []
        for csv_path, csv_type in zip(csv_paths, _CSV_TYPES):
            df_csv = pd.read_csv(csv_path)
            for _, row in df_csv.iterrows():
                new_texts.append(
                    serialize_record(row, csv_name=csv_type,
                                     use_block_tokens=use_block_tokens, skip_null=skip_null)
                )
        if len(new_texts) != len(records_df):
            raise ValueError(
                f"Mismatch: CSVs producen {len(new_texts)} registros, "
                f"records_interim tiene {len(records_df)}. ¿Los CSVs cambiaron?"
            )
        records_df = records_df.copy()
        records_df["text"] = new_texts
        print(f"✓ Re-serializado text con use_block_tokens={use_block_tokens}, skip_null={skip_null}")

    # Si el xlsx ya tiene columnas entity_id_* (iteración previa de finalize), las
    # dropeamos para no duplicarlas al regenerar. Incluye 'entity_id' (singular) por
    # compatibilidad con xlsx de iteraciones tempranas.
    for col in ("entity_id", "entity_id_a", "entity_id_b"):
        if col in review_df.columns:
            review_df = review_df.drop(columns=[col])

    # Cuando todas las celdas de 'decision' están vacías, pandas infiere la columna como
    # float64 (NaN es un float) y rechaza la escritura posterior de strings ('match',
    # 'no_match') con LossySetitemError. Forzar object dtype para evitarlo.
    review_df["decision"] = review_df["decision"].astype(object)

    # Aplicar regla de precedencia entre 'decision' (manual) y 'criterio' (pipeline):
    #   decision == 'match'      → par positivo (incluso si criterio=no_confirmado)
    #   decision == 'no_match'   → par excluido del union-find (incluso si criterio=llave_exacta)
    #   decision vacío/NaN       → usa el criterio del pipeline
    decision = review_df["decision"]
    criterio = review_df["criterio"]

    is_match    = decision == "match"
    is_no_match = decision == "no_match"
    is_blank    = decision.isna() | (decision == "")

    positive_from_criterio = is_blank & criterio.isin(POSITIVE_CRITERIA)
    positive_mask = (is_match | positive_from_criterio) & (~is_no_match)

    n_manual_match = int(is_match.sum())
    n_manual_no_match = int(is_no_match.sum())
    if n_manual_match or n_manual_no_match:
        print(f"✓ Decisiones manuales aplicadas: {n_manual_match} match, {n_manual_no_match} no_match")

    positive_pairs = list(zip(
        review_df.loc[positive_mask, "record_id_a"].astype(int),
        review_df.loc[positive_mask, "record_id_b"].astype(int),
    ))

    # Intra-source grouping: registros duplicados dentro del mismo CSV que comparten
    # (exp_int, nombre_norm) no aparecen en pairs_classified (fueron deduplicados en build_pairs_df).
    # Se conectan aquí para que reciban el mismo entity_id que su representante.
    intra_pairs = []
    valid_meta = records_df[records_df["exp_int"].notna() & (records_df["nombre_norm"] != "")]
    for _, group in valid_meta.groupby(["source_db", "exp_int", "nombre_norm"]):
        ids = group["record_id"].tolist()
        for i in range(1, len(ids)):
            intra_pairs.append((ids[0], ids[i]))

    n_records = len(records_df)
    entity_ids = _union_find_entity_ids(n_records, positive_pairs + intra_pairs)
    records_df = records_df.copy()
    records_df["entity_id"] = entity_ids

    # Parquet final — mismo esquema que dataset.py v1
    df_output = records_df[["record_id", "source_db", "text", "entity_id"]]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_output.to_parquet(output_path, engine="pyarrow", index=False, compression="snappy")

    # Transiciones de criterio/decision para el estado final del pipeline:
    #   1. Auto-confirmados (llave_exacta, metrica_clasica) con decision vacío → decision=match.
    #      Refleja que el pipeline los promueve automáticamente; el revisor ya no decide sobre ellos.
    #   2. no_confirmado con decision filled (match o no_match) → criterio=revision_manual.
    #      La categoría no_confirmado se reserva para los que aún no han sido revisados.
    decision_filled = review_df["decision"].isin(["match", "no_match"])
    auto_confirmed_mask = review_df["criterio"].isin(POSITIVE_CRITERIA)
    no_confirmado_mask  = review_df["criterio"] == "no_confirmado"

    review_df.loc[auto_confirmed_mask & ~decision_filled, "decision"] = "match"
    review_df.loc[no_confirmado_mask & decision_filled, "criterio"] = "revision_manual"

    # Regenerar xlsx con entity_id_a y entity_id_b al final, preservando decision + dropdown.
    # Las dos columnas permiten auditar: filas con entity_id_a != entity_id_b son pares cuya
    # decisión (no_match o no_confirmado vacío) llevó a entidades distintas; filas con valores
    # iguales son pares que terminaron en la misma entidad (por union-find directo o transitivo).
    record_to_entity = dict(zip(records_df["record_id"], records_df["entity_id"]))
    review_df["entity_id_a"] = review_df["record_id_a"].astype(int).map(record_to_entity)
    review_df["entity_id_b"] = review_df["record_id_b"].astype(int).map(record_to_entity)
    _write_review_xlsx(review_df, xlsx_path)

    _print_finalize_summary(df_output, review_df)
    print(f"✓ Parquet final: {output_path}")
    print(f"✓ pairs_for_review.xlsx regenerado con entity_id: {xlsx_path}")

    return df_output


def _print_classify_summary(pairs_df: pd.DataFrame) -> None:
    counts = pairs_df["criterio"].value_counts()
    total = len(pairs_df)
    print(f"\n  Pares clasificados: {total:,}")
    for criterio in ["llave_exacta", "metrica_clasica", "no_confirmado"]:
        n = int(counts.get(criterio, 0))
        print(f"    {criterio:<22} {n:>6,}  ({100*n/total:.1f}%)")


def _print_finalize_summary(df_output: pd.DataFrame, review_df: pd.DataFrame) -> None:
    vc = df_output["entity_id"].value_counts()
    entity_sources = df_output.groupby("entity_id")["source_db"].apply(set)
    cross_db = int(entity_sources.apply(lambda s: len(s) * (len(s) - 1) // 2).sum())
    total_pares = len(review_df)

    n_llave_exacta    = int((review_df["criterio"] == "llave_exacta").sum())
    n_metrica_clasica = int((review_df["criterio"] == "metrica_clasica").sum())
    n_revision_manual = int((review_df["criterio"] == "revision_manual").sum())
    n_pendientes      = int((review_df["criterio"] == "no_confirmado").sum())

    n_match    = int((review_df["decision"] == "match").sum())
    n_no_match = int((review_df["decision"] == "no_match").sum())

    print(f"\n  Registros:                       {len(df_output):,}")
    print(f"  Entidades únicas:                {df_output['entity_id'].nunique():,}")
    print(f"  Pares positivos (in-batch):      {int((vc * (vc - 1) // 2).sum()):,}")
    print(f"  Pares confirmados cross-db:      {cross_db:,}  ← baseline v1: 9,855")
    print(f"\n  Por criterio (post-finalize):")
    print(f"    llave_exacta                   {n_llave_exacta:>6,}")
    print(f"    metrica_clasica                {n_metrica_clasica:>6,}")
    print(f"    revision_manual                {n_revision_manual:>6,}")
    print(f"    no_confirmado (sin revisar)    {n_pendientes:>6,}")
    print(f"\n  Por decision (veredicto final):")
    print(f"    match                          {n_match:>6,}")
    print(f"    no_match                       {n_no_match:>6,}")
    print(f"    vacío (sin revisar)            {n_pendientes:>6,}")
    print(f"    {'─'*40}")
    print(f"    total                          {total_pares:>6,}")

    if n_pendientes > 0:
        print(f"\n  ⚠ {n_pendientes} pares quedaron como no_confirmado sin decisión — pipeline incompleto.")
        print(f"    Edita la columna 'decision' en pairs_for_review.xlsx y vuelve a correr --step finalize.")
